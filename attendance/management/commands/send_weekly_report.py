"""
Management command: send_weekly_report

Generates a weekly attendance summary for the PREVIOUS Monday–Sunday period
and emails it as an Excel attachment to all active admins and program managers.

Also sends each manager an in-app + push notification linking to the admin
attendance page so they know the report has landed in their inbox.

Schedule: run once on Monday morning (handled by run_scheduler.sh).
Safe to re-run: a duplicate-send guard prevents sending more than once per week.
"""

import io
import datetime
import logging

from django.core.management.base import BaseCommand
from django.utils import timezone

logger = logging.getLogger(__name__)


class Command(BaseCommand):
    help = "Email the previous week's attendance summary to managers."

    def add_arguments(self, parser):
        parser.add_argument(
            "--force",
            action="store_true",
            help="Send even if a report was already sent this week (for testing).",
        )

    def handle(self, *args, **options):
        from accounts.models import User
        from attendance.models import Attendance, ProjectSite
        from communication.models import Notification

        now   = timezone.localtime()
        today = now.date()

        # ── Date range: previous Mon → Sun ───────────────────────────────────
        # today.weekday(): 0=Mon … 6=Sun
        # We always want the LAST full week regardless of which day we run.
        days_since_monday = today.weekday()  # 0 if today is Mon
        week_end   = today - datetime.timedelta(days=days_since_monday + 1)   # last Sun
        week_start = week_end - datetime.timedelta(days=6)                    # last Mon

        # ── Duplicate-send guard ─────────────────────────────────────────────
        guard_title = f"📊 Weekly attendance report — w/e {week_end.strftime('%d %b %Y')}"

        if not options.get("force"):
            already = Notification.objects.filter(
                title=guard_title,
                timestamp__date=today,
            ).exists()
            if already:
                self.stdout.write(f"  Report for w/e {week_end} already sent today — skipping.")
                return

        # ── Gather recipients ────────────────────────────────────────────────
        managers = list(
            User.objects.filter(
                is_active=True,
                role__in=["admin", "program_manager"],
                email__isnull=False,
            ).exclude(email="")
        )

        if not managers:
            self.stdout.write(self.style.WARNING("  No managers with email addresses — skipping."))
            return

        # ── Build per-user stats for the week ────────────────────────────────
        staff = User.objects.filter(
            is_active=True,
            role__in=["staff", "intern", "program_manager", "department_head", "admin"],
        ).select_related("department").order_by("first_name", "username")

        # Count working days in the range (Mon–Fri only)
        working_days = sum(
            1 for i in range((week_end - week_start).days + 1)
            if (week_start + datetime.timedelta(days=i)).weekday() < 5
        )

        records_qs = Attendance.objects.filter(
            check_in_time__date__gte=week_start,
            check_in_time__date__lte=week_end,
        ).select_related("user", "project_site")

        # Index records by user
        from collections import defaultdict
        by_user = defaultdict(list)
        for rec in records_qs:
            by_user[rec.user_id].append(rec)

        rows = []
        summary_lines = []  # for email body text
        for user in staff:
            recs = by_user.get(user.pk, [])
            days_present = len(recs)
            days_absent  = max(0, working_days - days_present)
            total_hours  = sum(float(r.total_hours or 0) for r in recs)
            lates        = sum(1 for r in recs if r.arrival_status == "late")
            early_outs   = sum(1 for r in recs if r.departure_status == "left_early")
            avg_hours    = round(total_hours / days_present, 2) if days_present else 0
            dept         = user.department.name if user.department else "—"

            rows.append([
                user.get_full_name() or user.username,
                user.email or "—",
                dept,
                user.get_role_display() if hasattr(user, "get_role_display") else user.role,
                days_present,
                days_absent,
                f"{total_hours:.1f}h",
                f"{avg_hours:.1f}h",
                lates,
                early_outs,
            ])

            if days_absent > 0 or lates > 0 or early_outs > 0:
                flags = []
                if days_absent: flags.append(f"{days_absent} absent")
                if lates:       flags.append(f"{lates} late arrival{'s' if lates!=1 else ''}")
                if early_outs:  flags.append(f"{early_outs} early departure{'s' if early_outs!=1 else ''}")
                summary_lines.append(
                    f"  • {user.get_full_name() or user.username}: {', '.join(flags)}"
                )

        headers = [
            "Staff Member", "Email", "Department", "Role",
            "Days Present", "Days Absent",
            "Total Hours", "Avg Hours/Day",
            "Late Arrivals", "Early Departures",
        ]

        # ── Generate Excel workbook in memory ────────────────────────────────
        excel_bytes = self._build_excel(
            headers, rows, week_start, week_end, working_days
        )

        # ── Send email to each manager ────────────────────────────────────────
        week_range_str = f"{week_start.strftime('%d %b')} – {week_end.strftime('%d %b %Y')}"
        subject = f"[SPH Portal] Weekly Attendance Report — {week_range_str}"

        issues_text = (
            "\n\nStaff requiring attention this week:\n" + "\n".join(summary_lines)
            if summary_lines
            else "\n\nNo attendance issues recorded this week. 🎉"
        )

        body = (
            f"Hello,\n\n"
            f"Please find attached the weekly attendance summary for "
            f"{week_range_str} ({working_days} working day{'s' if working_days!=1 else ''}).\n"
            f"\nTotal staff: {len(rows)}"
            f"\nPresent at least once: {sum(1 for r in rows if r[4] > 0)}"
            f"\nFull attendance (no absences): {sum(1 for r in rows if r[5] == 0)}"
            f"{issues_text}\n\n"
            f"View full attendance details: /attendance/admin/\n\n"
            f"— Swahilipot Hub Portal (automated report)"
        )

        from django.core.mail import EmailMessage
        fname = f"attendance-{week_start.strftime('%Y%m%d')}-{week_end.strftime('%Y%m%d')}.xlsx"

        sent_count = 0
        for mgr in managers:
            try:
                email = EmailMessage(
                    subject=subject,
                    body=body,
                    to=[mgr.email],
                )
                email.attach(fname, excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                email.send(fail_silently=False)
                sent_count += 1
                self.stdout.write(f"  ✓ Sent to {mgr.email}")
            except Exception as exc:
                logger.warning("Weekly report email failed for %s: %s", mgr.email, exc)
                self.stdout.write(self.style.ERROR(f"  ✗ Failed for {mgr.email}: {exc}"))

        # ── In-app + push notification to each manager ────────────────────────
        from core.notify import notify_managers, MEDIUM
        notify_managers(
            title=guard_title,
            message=(
                f"The weekly attendance report for {week_range_str} has been emailed to you. "
                f"{len(rows)} staff members covered, {working_days} working days."
                + (f" {len(summary_lines)} staff member(s) had attendance issues." if summary_lines else "")
            ),
            priority=MEDIUM,
            link="/attendance/admin/",
        )

        self.stdout.write(
            self.style.SUCCESS(
                f"\n  📊 Weekly report sent to {sent_count}/{len(managers)} manager(s) "
                f"for {week_range_str}."
            )
        )

    # ────────────────────────────────────────────────────────────────────────
    def _build_excel(self, headers, rows, week_start, week_end, working_days):
        """Build a styled Excel workbook and return bytes."""
        try:
            import openpyxl
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side
            )
            from openpyxl.utils import get_column_letter
        except ImportError:
            # Fallback: plain CSV bytes
            lines = [",".join(str(h) for h in headers)]
            for row in rows:
                lines.append(",".join(f'"{v}"' for v in row))
            return "\n".join(lines).encode()

        NAVY   = "1E3A5F"
        BLUE   = "1E40AF"
        GOLD   = "F59E0B"
        LIGHT  = "EFF6FF"
        STRIPE = "DBEAFE"
        RED    = "FEE2E2"
        GREEN  = "D1FAE5"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Weekly Report"
        ws.sheet_view.showGridLines = False

        n_cols   = len(headers)
        last_col = get_column_letter(n_cols)

        def fill(hex_):
            return PatternFill("solid", fgColor=hex_)

        def font(color="000000", bold=False, size=11, italic=False):
            return Font(color=color, bold=bold, size=size, italic=italic, name="Calibri")

        def border(color="BFDBFE"):
            s = Side(style="thin", color=color)
            return Border(left=s, right=s, top=s, bottom=s)

        # Row 1 — brand banner
        ws.merge_cells(f"A1:{last_col}1")
        ws["A1"] = "  Swahilipot Hub Portal"
        ws["A1"].font      = font("FFFFFF", bold=True, size=14)
        ws["A1"].fill      = fill(NAVY)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 32

        # Row 2 — report title
        week_range_str = f"{week_start.strftime('%d %b')} – {week_end.strftime('%d %b %Y')}"
        ws.merge_cells(f"A2:{last_col}2")
        ws["A2"] = f"  Weekly Attendance Report  |  {week_range_str}  |  {working_days} working days"
        ws["A2"].font      = font("DBEAFE", size=9, italic=True)
        ws["A2"].fill      = fill(BLUE)
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 20

        # Row 3 — blank spacer
        ws.row_dimensions[3].height = 8

        # Row 4 — column headers
        for col_idx, hdr in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_idx, value=hdr)
            cell.font      = font("FFFFFF", bold=True, size=10)
            cell.fill      = fill(BLUE)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = border(NAVY)
        ws.row_dimensions[4].height = 30

        # Data rows
        for row_idx, row in enumerate(rows, start=5):
            is_stripe = (row_idx % 2 == 0)
            days_absent   = row[5] if isinstance(row[5], int) else 0
            lates         = row[8] if isinstance(row[8], int) else 0
            early_outs    = row[9] if isinstance(row[9], int) else 0
            has_issue     = days_absent > 0 or lates > 0 or early_outs > 0
            row_fill      = fill(RED if has_issue else (STRIPE if is_stripe else "FFFFFF"))

            for col_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font      = font(size=10)
                cell.fill      = row_fill
                cell.border    = border()
                cell.alignment = Alignment(
                    horizontal="right" if col_idx > 4 else "left",
                    vertical="center",
                )
            ws.row_dimensions[row_idx].height = 18

        # Summary row
        summary_row = len(rows) + 5
        total_present = sum(r[4] for r in rows if isinstance(r[4], int))
        total_hours   = sum(float(r[6].replace("h","")) for r in rows if isinstance(r[6], str))
        total_lates   = sum(r[8] for r in rows if isinstance(r[8], int))
        total_early   = sum(r[9] for r in rows if isinstance(r[9], int))

        ws.cell(summary_row, 1, "TOTALS").font = font(bold=True, size=10)
        ws.cell(summary_row, 1).fill  = fill(GOLD)
        ws.cell(summary_row, 5, total_present).fill = fill(GOLD)
        ws.cell(summary_row, 5).font = font(bold=True)
        ws.cell(summary_row, 7, f"{total_hours:.1f}h").fill = fill(GOLD)
        ws.cell(summary_row, 7).font = font(bold=True)
        ws.cell(summary_row, 9, total_lates).fill = fill(GOLD)
        ws.cell(summary_row, 9).font = font(bold=True)
        ws.cell(summary_row, 10, total_early).fill = fill(GOLD)
        ws.cell(summary_row, 10).font = font(bold=True)
        for col_idx in range(1, n_cols + 1):
            ws.cell(summary_row, col_idx).border = border(NAVY)
        ws.row_dimensions[summary_row].height = 22

        # Column widths
        col_widths = [28, 30, 18, 18, 13, 13, 12, 14, 14, 16]
        for i, w in enumerate(col_widths[:n_cols], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Freeze header rows
        ws.freeze_panes = "A5"

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
